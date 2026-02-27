"""infer_schema — detect column types by sampling cell values.

Samples up to 50 non-empty cells per column and uses regex / heuristic
rules to classify each column as one of:
  currency, percent, date, email, url, id, categorical, number, text
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# Type-detection regexes
# ---------------------------------------------------------------------------

_RE_CURRENCY = re.compile(r"^[\$€R\$£¥]\s*[\d,]+\.?\d*$|^[\d,]+\.?\d*\s*[\$€R\$£¥]$")
_RE_PERCENT = re.compile(r"^\d+(\.\d+)?%$")
_RE_EMAIL = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")
_RE_URL = re.compile(r"^https?://\S+$", re.IGNORECASE)
_RE_DATE_SLASH = re.compile(r"^\d{1,4}[/\-\.]\d{1,2}[/\-\.]\d{1,4}$")
_RE_INTEGER = re.compile(r"^-?\d+$")
_RE_NUMBER = re.compile(r"^-?[\d,]+\.?\d*$")

_SAMPLE_SIZE = 50
_CATEGORICAL_UNIQUE_THRESHOLD = 20
_CATEGORICAL_MIN_ROWS = 100


def _classify_value(val: Any) -> str:
    """Return a type tag for a single cell value."""
    if val is None:
        return "empty"

    if isinstance(val, datetime):
        return "date"

    if isinstance(val, bool):
        return "text"  # bools are not numeric in spreadsheet context

    if isinstance(val, (int, float)):
        # 0-1 float heuristic for percent
        if isinstance(val, float) and 0.0 <= val <= 1.0:
            return "percent_candidate"
        # Excel serial date: roughly 1..2958465 (1900-01-01 .. 9999-12-31)
        if isinstance(val, (int, float)) and 1 <= val <= 2958465 and isinstance(val, int):
            return "date_serial_candidate"
        return "number"

    s = str(val).strip()
    if not s:
        return "empty"

    if _RE_CURRENCY.match(s):
        return "currency"
    if _RE_PERCENT.match(s):
        return "percent"
    if _RE_EMAIL.match(s):
        return "email"
    if _RE_URL.match(s):
        return "url"
    if _RE_DATE_SLASH.match(s):
        return "date"
    if _RE_NUMBER.match(s):
        return "number"
    return "text"


def _decide_column_type(tags: list[str], unique_count: int, total_rows: int) -> str:
    """Given collected type-tags for a column, decide the dominant type."""
    if not tags:
        return "text"

    # Count occurrences
    counts: dict[str, int] = {}
    for t in tags:
        counts[t] = counts.get(t, 0) + 1

    # Merge candidates
    if "percent_candidate" in counts:
        counts["percent"] = counts.get("percent", 0) + counts.pop("percent_candidate")
    if "date_serial_candidate" in counts:
        counts["number"] = counts.get("number", 0) + counts.pop("date_serial_candidate")

    counts.pop("empty", None)

    if not counts:
        return "text"

    dominant = max(counts, key=lambda k: counts[k])

    # ID heuristic: sequential unique integers
    if dominant == "number" and unique_count == len(tags):
        return "id"

    # Categorical heuristic
    if dominant == "text" and total_rows >= _CATEGORICAL_MIN_ROWS and unique_count < _CATEGORICAL_UNIQUE_THRESHOLD:
        return "categorical"

    return dominant


def run(
    *,
    workbook: Any | None = None,
    file_bytes: bytes | None = None,
    sheet_name: str | None = None,
    **_kwargs: Any,
) -> dict:
    """Entry-point for the infer_schema recipe."""
    wb: Any
    if workbook is not None:
        wb = workbook
    elif file_bytes is not None:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    else:
        return {"columns": []}

    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws is None:
        return {"columns": []}

    max_col = ws.max_column or 0
    max_row = ws.max_row or 0

    columns: list[dict] = []

    for c in range(1, max_col + 1):
        header = ws.cell(row=1, column=c).value
        header_str = str(header).strip() if header is not None else f"col_{c}"

        # Sample up to _SAMPLE_SIZE non-empty values (skip header row)
        samples: list[Any] = []
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                samples.append(val)
            if len(samples) >= _SAMPLE_SIZE:
                break

        tags = [_classify_value(v) for v in samples]
        unique_count = len(set(str(v) for v in samples))
        col_type = _decide_column_type(tags, unique_count, max_row - 1)

        columns.append(
            {
                "column": header_str,
                "index": c,
                "inferred_type": col_type,
                "sample_size": len(samples),
                "unique_count": unique_count,
            }
        )

    return {"columns": columns}
