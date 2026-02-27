"""trim_whitespace — strip leading/trailing whitespace from string cells.

Operates on an openpyxl Workbook or raw file bytes and returns the count
of cells that were actually modified.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


def run(
    *,
    workbook: Any | None = None,
    file_bytes: bytes | None = None,
    sheet_name: str | None = None,
    columns: list[int | str] | None = None,
    **_kwargs: Any,
) -> dict:
    """Entry-point for the trim_whitespace recipe.

    Parameters
    ----------
    workbook:
        An openpyxl ``Workbook`` instance.
    file_bytes:
        Raw Excel file bytes (alternative to *workbook*).
    sheet_name:
        Target sheet name (default: active sheet).
    columns:
        List of 1-based column indices **or** column letters to process.
        ``None`` means all columns.

    Returns
    -------
    dict with ``trimmed_count``.
    """
    wb: Any
    if workbook is not None:
        wb = workbook
    elif file_bytes is not None:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    else:
        return {"trimmed_count": 0}

    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws is None:
        return {"trimmed_count": 0}

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Resolve target column indices (1-based)
    target_cols: list[int]
    if columns is not None:
        target_cols = []
        for col in columns:
            if isinstance(col, int):
                target_cols.append(col)
            elif isinstance(col, str) and col.isalpha():
                # Convert letter(s) to column index
                idx = 0
                for ch in col.upper():
                    idx = idx * 26 + (ord(ch) - ord("A") + 1)
                target_cols.append(idx)
            else:
                # Try int conversion
                try:
                    target_cols.append(int(col))
                except (ValueError, TypeError):
                    pass
    else:
        target_cols = list(range(1, max_col + 1))

    trimmed_count = 0

    for r in range(1, max_row + 1):
        for c in target_cols:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, str):
                stripped = cell.value.strip()
                if stripped != cell.value:
                    cell.value = stripped
                    trimmed_count += 1

    return {"trimmed_count": trimmed_count}
