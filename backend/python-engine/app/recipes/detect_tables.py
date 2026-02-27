"""detect_tables — scan a workbook for contiguous non-empty rectangular regions.

Accepts either an openpyxl ``Workbook`` object (via ``workbook`` kwarg) or
raw ``bytes`` (via ``file_bytes`` kwarg).

Returns
-------
dict  with key ``tables`` — a list of dicts:
    ``{"sheet": str, "range": str, "rows": int, "cols": int, "has_header": bool}``
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


def _looks_like_header(row: tuple) -> bool:
    """Heuristic: a row is a header when every cell is a non-empty string."""
    if not row:
        return False
    return all(isinstance(c, str) and c.strip() for c in row)


def _find_regions(ws: Any) -> list[dict]:
    """Find contiguous non-empty rectangular regions in *ws*."""
    regions: list[dict] = []

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return regions

    # Build a boolean grid of non-empty cells
    grid: list[list[bool]] = []
    for r in range(1, max_row + 1):
        row_flags: list[bool] = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            row_flags.append(val is not None and str(val).strip() != "")
        grid.append(row_flags)

    visited = [[False] * max_col for _ in range(max_row)]

    for r in range(max_row):
        for c in range(max_col):
            if visited[r][c] or not grid[r][c]:
                continue

            # Expand right as far as contiguous non-empty cells on this row
            c_end = c
            while c_end + 1 < max_col and grid[r][c_end + 1]:
                c_end += 1

            # Expand down: keep the bounding box width and extend rows as long
            # as at least half the cells in the row-slice are non-empty
            r_end = r
            width = c_end - c + 1
            for rr in range(r + 1, max_row):
                filled = sum(1 for cc in range(c, c_end + 1) if grid[rr][cc])
                if filled >= max(1, width // 2):
                    r_end = rr
                else:
                    break

            # Only report regions with at least 2 rows and 2 columns
            height = r_end - r + 1
            if height >= 2 and width >= 2:
                first_row_values = tuple(
                    ws.cell(row=r + 1, column=cc + 1).value for cc in range(c, c_end + 1)
                )
                has_header = _looks_like_header(first_row_values)

                start_cell = f"{get_column_letter(c + 1)}{r + 1}"
                end_cell = f"{get_column_letter(c_end + 1)}{r_end + 1}"

                regions.append(
                    {
                        "sheet": ws.title,
                        "range": f"{start_cell}:{end_cell}",
                        "rows": height,
                        "cols": width,
                        "has_header": has_header,
                    }
                )

            # Mark visited
            for rr in range(r, r_end + 1):
                for cc in range(c, c_end + 1):
                    visited[rr][cc] = True

    return regions


def run(
    *,
    workbook: Any | None = None,
    file_bytes: bytes | None = None,
    **_kwargs: Any,
) -> dict:
    """Entry-point for the detect_tables recipe."""
    wb: Any
    if workbook is not None:
        wb = workbook
    elif file_bytes is not None:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    else:
        # Return empty when no workbook provided (e.g. dry-run)
        return {"tables": []}

    tables: list[dict] = []
    for ws in wb.worksheets:
        tables.extend(_find_regions(ws))

    return {"tables": tables}
