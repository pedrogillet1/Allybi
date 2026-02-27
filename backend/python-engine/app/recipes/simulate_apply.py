"""simulate_apply — dry-run application of ops on a workbook copy.

Takes a workbook (openpyxl Workbook or raw bytes) and a list of cell-level
mutations, applies them to a **copy** of the workbook, and returns a diff
of changed cell values.
"""

from __future__ import annotations

import copy
import io
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


def _cell_key(sheet: str, row: int, col: int) -> str:
    return f"{sheet}!{get_column_letter(col)}{row}"


def _snapshot(wb: Any) -> dict[str, Any]:
    """Build a dict of ``{sheet!CellRef: value}`` for every non-empty cell."""
    snap: dict[str, Any] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    snap[_cell_key(ws.title, cell.row, cell.column)] = cell.value
    return snap


def _apply_mutations(wb: Any, mutations: list[dict]) -> None:
    """Apply a list of ``{"sheet", "cell", "value"}`` mutations to *wb*."""
    for mut in mutations:
        sheet_name = mut.get("sheet")
        cell_ref = mut.get("cell")
        value = mut.get("value")

        if not sheet_name or not cell_ref:
            continue

        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ws[cell_ref] = value


def run(
    *,
    workbook: Any | None = None,
    file_bytes: bytes | None = None,
    mutations: list[dict] | None = None,
    **_kwargs: Any,
) -> dict:
    """Entry-point for the simulate_apply recipe.

    Parameters
    ----------
    workbook:
        An openpyxl ``Workbook`` to operate on (will be deep-copied).
    file_bytes:
        Raw Excel file bytes (alternative to *workbook*).
    mutations:
        List of ``{"sheet": str, "cell": str, "value": Any}`` dicts.

    Returns
    -------
    dict with ``changed_cells`` count and ``preview`` list of diffs.
    """
    if mutations is None or len(mutations) == 0:
        return {"changed_cells": 0, "preview": []}

    wb_original: Any
    if workbook is not None:
        wb_original = workbook
    elif file_bytes is not None:
        wb_original = openpyxl.load_workbook(io.BytesIO(file_bytes))
    else:
        return {"changed_cells": 0, "preview": []}

    before = _snapshot(wb_original)

    # Deep-copy so the original is untouched
    wb_copy = copy.deepcopy(wb_original)
    _apply_mutations(wb_copy, mutations)

    after = _snapshot(wb_copy)

    # Diff
    all_keys = set(before.keys()) | set(after.keys())
    preview: list[dict] = []

    for key in sorted(all_keys):
        old_val = before.get(key)
        new_val = after.get(key)
        if old_val != new_val:
            preview.append(
                {
                    "cell": key,
                    "before": old_val,
                    "after": new_val,
                }
            )

    return {"changed_cells": len(preview), "preview": preview}
