"""Tests for the detect_tables recipe."""

from __future__ import annotations

import openpyxl
import pytest

from app.recipes.detect_tables import run


def _make_workbook_with_table(
    *,
    start_row: int = 1,
    start_col: int = 1,
    num_rows: int = 5,
    num_cols: int = 3,
    sheet_name: str = "Sheet1",
    headers: list[str] | None = None,
) -> openpyxl.Workbook:
    """Create an in-memory workbook with data written to a known range."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name

    if headers is None:
        headers = [f"Col{c}" for c in range(1, num_cols + 1)]

    # Write header row
    for c_offset, header in enumerate(headers):
        ws.cell(row=start_row, column=start_col + c_offset, value=header)

    # Write data rows
    for r_offset in range(1, num_rows):
        for c_offset in range(num_cols):
            ws.cell(
                row=start_row + r_offset,
                column=start_col + c_offset,
                value=f"val_{r_offset}_{c_offset}",
            )

    return wb


class TestDetectTables:
    def test_finds_single_table(self):
        wb = _make_workbook_with_table(num_rows=6, num_cols=4)
        result = run(workbook=wb)
        tables = result["tables"]
        assert len(tables) == 1
        table = tables[0]
        assert table["sheet"] == "Sheet1"
        assert table["rows"] == 6
        assert table["cols"] == 4
        assert table["has_header"] is True

    def test_finds_table_at_offset(self):
        wb = _make_workbook_with_table(start_row=3, start_col=2, num_rows=4, num_cols=3)
        result = run(workbook=wb)
        tables = result["tables"]
        assert len(tables) == 1
        assert tables[0]["rows"] == 4
        assert tables[0]["cols"] == 3
        # Range should start at B3
        assert tables[0]["range"].startswith("B3")

    def test_empty_workbook_returns_no_tables(self):
        wb = openpyxl.Workbook()
        result = run(workbook=wb)
        assert result["tables"] == []

    def test_no_workbook_returns_empty(self):
        result = run()
        assert result == {"tables": []}

    def test_numeric_header_is_not_header(self):
        """A row of numbers should not be detected as a header."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        # Fill a 3x3 grid of numbers
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(row=r, column=c, value=r * 10 + c)
        result = run(workbook=wb)
        if result["tables"]:
            assert result["tables"][0]["has_header"] is False

    def test_multiple_sheets(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Data"
        for r in range(1, 5):
            for c in range(1, 4):
                ws1.cell(row=r, column=c, value=f"d{r}{c}" if r > 1 else f"Header{c}")

        ws2 = wb.create_sheet("Other")
        for r in range(1, 4):
            for c in range(1, 3):
                ws2.cell(row=r, column=c, value=f"o{r}{c}" if r > 1 else f"H{c}")

        result = run(workbook=wb)
        sheets_found = {t["sheet"] for t in result["tables"]}
        assert "Data" in sheets_found
        assert "Other" in sheets_found
